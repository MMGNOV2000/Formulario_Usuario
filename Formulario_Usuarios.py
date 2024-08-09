import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os

nombre_archivo = 'datos.xlsx'

if os.path.exists(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Edad", "Email", "Telefono", "Direccion"])
# Se creara el libro de Excel

# wb.save('datos.xlsx')


ventana = tk.Tk()
ventana.geometry("252x252")

ventana.title("Formulario de Entrada de Datos") # creamos el titulo del la centana con el nombre formulario 
ventana.configure(bg= "#4B6587")
label_style = {"bg": "#4B6587", "fg": "white"}
entry_style = {"bg": '#D3D3D3', "fg": "black"}

def guardar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    email = entry_email.get()
    telefonno = entry_telefono.get()
    direccion = entry_direccion.get()
    
    if not nombre or not edad or not email or not telefonno or not direccion:
        messagebox.showwarning("Advertencia", "todos los campos son obligatorios")
        return
    try: 
        edad = int(edad)
        telefonno = int(telefonno)
    except ValueError:
        messagebox.showwarning("Advertencia", "Edad y Telefono deben ser numeros")
        return
    if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
        messagebox.showwarning("Advertencia", "El correo electrónico no es válido")
        return
    
    ws.append([nombre, edad, email, telefonno, direccion])
    wb.save(nombre_archivo)
    messagebox.showinfo("Informacion", "Datos guardados con exito")


    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)

# se creacion de las etiqueta
label_nombre = tk.Label(ventana, text="Nombre", font=("Arial", 12), **label_style) # creamos la etiqueta con el texto principal "Nombre"
label_nombre.grid(row=0, column=0, padx=10, pady= 5)  # Colocamos la etiqueta

entry_nombre = tk.Entry(ventana, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady= 5)

label_edad = tk.Label(ventana, text="Edad", font=("Arial", 12), **label_style)
label_edad.grid(row= 1, column= 0, padx= 15, pady= 6)

entry_edad = tk.Entry(ventana, **entry_style)
entry_edad.grid(row= 1, column= 1, padx= 15, pady= 6)

label_email = tk.Label(ventana, text="Email", font=("Arial", 12), **label_style)
label_email.grid(row= 2, column= 0, padx= 15, pady= 6)

entry_email = tk.Entry(ventana, **entry_style)
entry_email.grid(row= 2, column= 1, padx= 15, pady= 6 )

label_telefono = tk.Label(ventana, text="Telefono", font=("Arial", 12), **label_style)
label_telefono.grid(row= 3, column= 0, padx= 15, pady= 6)

entry_telefono = tk.Entry(ventana, **entry_style)
entry_telefono.grid(row=3, column=1, padx=15, pady=6)

label_direccion = tk.Label(ventana, text="Direccion", font=("Arial", 12), **label_style)
label_direccion.grid(row= 4, column= 0, padx= 15, pady= 6)

entry_direccion = tk.Entry(ventana, **entry_style)
entry_direccion.grid(row= 4, column= 1, padx= 15, pady= 6)

boton_guardar = tk.Button(ventana, text="Guardar", command= guardar_datos,bg= "#4B6587",fg= "white",  font=("Arial", 12))
boton_guardar.grid(row= 5, column= 1, padx= 15, pady= 6)


ventana.mainloop()
